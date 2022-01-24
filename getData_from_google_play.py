#!/usr/bin/python3
#-*- coding: UTF-8 -*-


import sys
import imp
import time
import urllib
#import urllib2
import urllib.request
import requests
import numpy as np
import xlrd
from xml.dom.minidom import parse
import os, sys,time,os.path,shutil,argparse
from bs4 import BeautifulSoup
from openpyxl import Workbook

#reload(sys)
imp.reload(sys)



#Some User Agents
hds=[{'User-Agent':'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6'},\
{'User-Agent':'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.12 Safari/535.11'},\
{'User-Agent': 'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0)'}]
#帮助信息
parser = argparse.ArgumentParser(description='create prebuild jar')
parser.add_argument('--path', help='excel path')
args = parser.parse_args()
excelpath=args.path
fromLine=1
toLine=50

def book_spider(book_tag):
    page_num=0;
    book_list=[]
    try_times=0
    listNum=1;
    findNum=0
    
    while(1):
        if(findNum==1):
                findNum=0
                break;
        url='https://apkcombo.net/search?q='+book_tag
        #print(url)
        time.sleep(np.random.rand())
        
        #Last Version
        try:
            req = urllib.request.Request(url, headers=hds[page_num%len(hds)])
            source_code = urllib.request.urlopen(req).read()
            plain_text=str(source_code)   
        except Exception as e:
            print(e)
            continue
  
        ##Previous Version, IP is easy to be Forbidden
        #source_code = requests.get(url) 
        #plain_text = source_code.text  
        
        soup = BeautifulSoup(plain_text)
        list_soup = soup.find('div', {'class': 'list-template lists'})
        
        try_times+=1;
        if list_soup==None and try_times<5:
            continue
        elif list_soup==None or len(list_soup)<=1:
            break # Break when no informatoin got after 200 times requesting
        print('111')
        
        for book_info in list_soup.findAll('div', {'class': 'list'}):           
            if(listNum>3):
                listNum=1;
                break;
            findNum=1
            try_times=0
            print('222')
            #print(book_info)
            title=book_info.a.attrs['href']
            #print(title)
            if(book_tag in title.split('/')):
                print('---find package path--')
                print(title)
                try:
                    req = urllib.request.Request('https://apkcombo.net/'+title, headers=hds[page_num%len(hds)])
                    source_code = urllib.request.urlopen(req).read()
                    plain_text=str(source_code)
                    soup = BeautifulSoup(plain_text)
                    list_soup = soup.find('div', {'class': 'item category'})
                    print(list_soup.a.attrs['href'])  
                    book_list.append([book_tag,list_soup.a.attrs['href']])
                except Exception as e:
                    print(e) 
                break
            listNum+=1
             #set 0 when got valid information
        page_num+=1
    return book_list


def get_people_num(url):
    #url='http://book.douban.com/subject/6082808/?from=tag_all' # For Test
    try:
        req = urllib.request.Request(url, headers=hds[np.random.randint(0,len(hds))])
        source_code = urllib.request.urlopen(req).read()
        plain_text=str(source_code)   
    except Exception as e:
        print(e)
    soup = BeautifulSoup(plain_text)
    people_num=soup.find('div',{'class':'rating_sum'}).findAll('span')[1].string.strip()
    return people_num


def do_spider(book_tag_lists):
    num_count=1
    book_lists=[]
    for book_tag in book_tag_lists:
        num_count+=1
        if(book_tag=='包名'):
             continue
        if(num_count>toLine):
             break
        print('+++++begin search+++++')
        print(num_count)
        print('https://apkcombo.net/search?q='+book_tag)
        book_list=book_spider(book_tag)
        print('*****end search*****\n')
        book_list=sorted(book_list,key=lambda x:x[1],reverse=True)
        if (len(book_list)==0):
             book_list.append([book_tag,'找不到'])
        book_lists.append(book_list)
    return book_lists


def print_book_lists_excel(book_lists,book_tag_lists):
    wb=Workbook(write_only=True)
    print(book_lists)
    ws=['package_categry']
    '''
    for i in range(len(book_tag_lists)):
        ws.append(wb.create_sheet(title=book_tag_lists[i].encode("utf-8").decode("utf-8"))) #utf8->unicode
    '''
    for i in range(len(book_tag_lists)): 
        ws.append(['包名','分类'])
        count=1
        for bl in book_lists[i]:
            ws.append([count,bl[0],bl[1]])
            count+=1
    save_path='book_list'
    for i in range(len(book_tag_lists)):
        save_path+=('-'+book_tag_lists[i].encode("utf-8").decode("utf-8"))
    save_path+='.xlsx'
    wb.save(save_path)
    
def print_book_lists_excel123(book_lists,book_tag_lists):
    wb=Workbook(write_only=True)
    book_names=['savePath']
    ws=[]
    for i in range(len(book_names)):
        ws.append(wb.create_sheet(title=book_names[i].encode("utf-8").decode("utf-8"))) #utf8->unicode
    ws[0].append(['序号','包名','类型'])
    for i in range(len(book_lists)): 
        count=1
        for bl in book_lists[i]:
            ws[0].append([count,bl[0],bl[1]])
            print('wirte excel')
            count+=1
    save_path='book_list'
    for i in range(len(book_names)):
        save_path+=('-'+book_names[i].encode("utf-8").decode("utf-8"))
    save_path+='.xlsx'
    wb.save(save_path)


def read_xml(in_path):
    '''读取并解析xml文件
       in_path: xml路径
       return: ElementTree'''
    tree = ElementTree()
    tree.parse(in_path)
    return tree


if __name__=='__main__':
    data=xlrd.open_workbook(excelpath)
    table_index=data.sheet_by_index(0)
    book_tag_lists=table_index.col_values(0)
    print(book_tag_lists)
    #book_tag_lists = ['com.google.android.deskclock','com.google.android.gms','com.lenovotab.camera','com.android.systemui','com.google.android.gm','com.appolo13.stickmandrawanimation']
    book_lists=do_spider(book_tag_lists)
    print_book_lists_excel123(book_lists,book_tag_lists)
    
