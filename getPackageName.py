#!/usr/bin/python3
# -*- coding: UTF-8 -*-

from xml.dom.minidom import parse
import xml.dom.minidom
import pandas as pd
from openpyxl import Workbook


#xmlPath= '/home/zhangxu/下载/114/resources/res/xml/compatible_rule.xml'
#xmlPath= '/media/zhangxu/disk/LenovoPM/app/src/main/res/xml/compatible_rule.xml'
#xmlPath= '/home/zhangxu/下载/test/compatible_rule.xml'
xmlPath= '/media/zhangxu/disk/aosp-11-GSI/vendor/lenovo/productivity/service/res/xml/compatible_rule.xml'
path = '/home/zhangxu/下载/'
xmlPath1 = '/media/zhangxu/disk/下载/pcmode_app_window_config.xml'
packageNames = []

# 使用minidom解析器打开 XML 文档
#DOMTree = xml.dom.minidom.parse(path+ 'compatible_rule_lgsi.xml')
DOMTree = xml.dom.minidom.parse(xmlPath1)
collection = DOMTree.documentElement
# 在集合中获取所有package
packages = collection.getElementsByTagName('App')


# file_handle=open('/home/zhangxu/下载/packageName.xlsx',mode='w')
file_handle = '/home/zhangxu/下载/packageName.xlsx'
wb = Workbook()
ws = wb.active
row_num = 2
for package in packages:
    appName=package.getAttribute('name')
    print ('\033[1;35m  appName:  ' + appName)
    ws.cell(row=row_num, column=1, value=appName)

    packageName=package.getAttribute('packageName')
    print ('\033[1;35m  包名:  ' + packageName)
    packageNames.append(packageName)
    ws.cell(row=row_num, column=2, value=packageName)
    
    startUp = package.getElementsByTagName('StartUp')
    if not(startUp):
        print ('\033[1;35m  startUp == null ')
        ws.cell(row=row_num, column=3, value=" NULL")
    else:
        print ('\033[1;35m  startUp:  ' + startUp.item(0).firstChild.data)
        startUp = startUp.item(0).firstChild.data
        if startUp == 'fullscreen':
            print ('\033[1;35m  startUp:  ' + startUp)
            ws.cell(row=row_num, column=3, value=startUp)
        else:
            print ('\033[1;35m  startUp:  ' + startUp)
            ws.cell(row=row_num, column=4, value=startUp) 

    
    unsupported = package.getElementsByTagName('Unsupported')
    if not(unsupported):
        # print ('\033[1;35m  startUp == null ')
        ws.cell(row=row_num, column=5, value="NULL")
    else:
        for node in unsupported:
            str1 = node.firstChild.data
            if str1 == 'toggleIcon':
                print ('\033[1;35m  unsupported ' + str1)
                ws.cell(row=row_num, column=5, value=node.firstChild.data)
            if str1 == 'resize':
                print ('\033[1;35m  unsupported ' + str1)
                ws.cell(row=row_num, column=6, value=node.firstChild.data)
            if str1 == 'kill':
                print ('\033[1;35m  unsupported ' + str1)
                ws.cell(row=row_num, column=7, value=node.firstChild.data)
            if str1 == 'relaunch':
                print ('\033[1;35m  unsupported ' + str1)
                ws.cell(row=row_num, column=8, value=node.firstChild.data)
            if str1 == 'fullrelaunch':
                print ('\033[1;35m  unsupported ' + str1)
                ws.cell(row=row_num, column=9, value=node.firstChild.data)
           




     # 增加行号
    row_num += 1 

wb.save(file_handle)
print(f"Data written to {file_handle}.")
	
# file_handle.close()
print ('\033[1;35mcompatible_rule文件中配置包名数量:' + str(len(packages)))

#read_excel()用来读取excel文件，记得加文件后缀
# data = pd.read_excel(path + 'P12 需要删除的应用.xlsx',sheet_name = 1)
# for index, value in data.iterrows():
#     pag = value.values
#     packageName = pag[2]
#     print (packageName)
#     for parent_node in packages:
#         if parent_node.getAttribute("app:packageName") == packageName:
#             collection.removeChild(parent_node)

#      #i = packageNames.index(packageName)
#      #xmlDoc=loadXMLDoc(xmlPath)
#      #y=DOMTree.getElementsByTagName("package")[i]
#      #DOMTree.documentElement.removeChild(y)
#      #print (i)
     
try:
     with open(xmlPath,'w') as fh:
        DOMTree.writexml(fh,indent='',addindent='\t',newl='\n',encoding='UTF-8')
        print('OK')
except Exception as err:
    print('错误：{err}'.format(err=err))
